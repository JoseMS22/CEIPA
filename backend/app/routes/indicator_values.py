from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
    status,
    UploadFile,
    File,
)
from sqlalchemy.orm import Session

from app.db import get_db
from app.schemas.indicator_value import (
    IndicatorValueCreate,
    IndicatorValueUpdate,
    IndicatorValueOut,
    PaginatedIndicatorValues,
)
from app.repositories import indicator_value_repo as repo
from .auth import require_admin, get_current_user, require_admin_or_analyst
from app.core.normalization import NormalizationError

# 👇 Ajusta estas rutas si tus modelos están en otro lado
from app.models.country import Country
from app.models.indicator import Indicator

# --------- extras para el Excel ----------
from openpyxl import load_workbook
from io import BytesIO
import unicodedata
import re

router = APIRouter(prefix="/indicator-values", tags=["IndicatorValues"])


# ================== ENDPOINTS EXISTENTES ==================

@router.get("", response_model=PaginatedIndicatorValues)
def list_indicator_values(
    scenario_id: int | None = Query(None),
    country_id: int | None = Query(None),
    indicator_id: int | None = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
):
    return repo.list_values(db, scenario_id, country_id, indicator_id, page, limit)


@router.post(
    "",
    response_model=IndicatorValueOut,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_admin_or_analyst)],
)
def upsert_indicator_value(
    payload: IndicatorValueCreate,
    db: Session = Depends(get_db),
    current=Depends(get_current_user),
):
    try:
        return repo.upsert_value(
            db, payload, user_id=current.id if current else None
        )
    except NormalizationError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.patch(
    "/{value_id}",
    response_model=IndicatorValueOut,
    dependencies=[Depends(require_admin_or_analyst)],
)
def update_indicator_value(
    value_id: int,
    payload: IndicatorValueUpdate,
    db: Session = Depends(get_db),
):
    iv = repo.get_by_id(db, value_id)
    if not iv:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    try:
        return repo.update_value(db, iv, payload)
    except NormalizationError as e:
        raise HTTPException(status_code=422, detail=str(e))


@router.delete(
    "/{value_id}",
    status_code=204,
    dependencies=[Depends(require_admin_or_analyst)],
)
def delete_indicator_value(value_id: int, db: Session = Depends(get_db)):
    iv = repo.get_by_id(db, value_id)
    if not iv:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    repo.delete_value(db, iv)
    return None


# ================== HELPERS PARA EXCEL ==================

def normalize_text(s: str) -> str:
    """
    Quita tildes, pasa a minúsculas y compacta espacios.
    Sirve para comparar 'Alemania', 'ALEMANIA', 'alemánia' como lo mismo.
    """
    if not s:
        return ""
    s = s.strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.lower()
    s = re.sub(r"\s+", " ", s)
    return s


def is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def is_text(value) -> bool:
    return isinstance(value, str) and value.strip() != ""


def build_country_map(db: Session):
    """
    Construye un dict para buscar países por:
    - iso2
    - iso3
    - name_es
    - name_en
    Usando normalize_text.
    """
    countries = db.query(Country).all()
    country_map: dict[str, Country] = {}
    ambiguous: set[str] = set()

    for c in countries:
        keys = [c.iso2, c.iso3, c.name_es, c.name_en]
        for key in keys:
            if not key:
                continue
            norm = normalize_text(str(key))
            if norm in country_map and country_map[norm].id != c.id:
                ambiguous.add(norm)
            else:
                country_map[norm] = c

    return country_map, ambiguous


def build_indicator_map(db: Session):
    """
    Construye un dict para buscar indicadores por nombre (indicator.name).
    """
    indicators = db.query(Indicator).all()
    indicator_map: dict[str, Indicator] = {}
    ambiguous: set[str] = set()

    for ind in indicators:
        norm = normalize_text(ind.name)
        if norm in indicator_map and indicator_map[norm].id != ind.id:
            ambiguous.add(norm)
        else:
            indicator_map[norm] = ind

    return indicator_map, ambiguous


def detect_headers(ws):
    """
    Detecta automáticamente:
    - header_row: fila donde están los nombres de los indicadores
    - header_col: columna donde están los nombres de los países

    Estrategia sencilla:
    - Fila con al menos 2 celdas de texto -> encabezado de columnas (indicadores)
    - Columna con al menos 2 celdas de texto -> encabezado de filas (países)
    """
    max_row = ws.max_row
    max_col = ws.max_column

    header_row_idx = None
    header_col_idx = None

    # Buscar fila de encabezados (indicadores)
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        text_cells = [v for v in row_vals if is_text(v)]
        if len(text_cells) >= 2:
            header_row_idx = r
            break

    # Buscar columna de encabezados (países)
    for c in range(1, max_col + 1):
        col_vals = [ws.cell(row=r, column=c).value for r in range(1, max_row + 1)]
        text_cells = [v for v in col_vals if is_text(v)]
        if len(text_cells) >= 2:
            header_col_idx = c
            break

    if header_row_idx is None or header_col_idx is None:
        raise ValueError(
            "No se pudo detectar una fila de encabezados de indicadores y una columna de países."
        )

    return header_row_idx, header_col_idx


# ================== NUEVO ENDPOINT: IMPORTAR EXCEL MATRIZ ==================

@router.post(
    "/import-matrix-excel",
    dependencies=[Depends(require_admin_or_analyst)],
)
async def import_indicator_values_matrix_excel(
    scenario_id: int = Query(..., ge=1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current=Depends(get_current_user),
):
    """
    Importa valores desde un Excel en forma de matriz:

        (col A)      (col B)        (col C)
      -----------------------------------------
      fila 1   |   (país / vacío)  | Indicador 1 | Indicador 2 ...
      fila 2   |   Colombia        |   3         |  36
      fila 3   |   México          |   7         |  80

    - Detecta automáticamente qué fila es encabezado de indicadores
      y qué columna es encabezado de países.
    - Ignora mayúsculas, tildes y espacios extras al comparar nombres.
    - Usa el repo.upsert_value para que la normalización funcione igual
      que cuando se hace manual.
    """

    # 1) Validar extensión
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser Excel (.xlsx, .xlsm o .xls).",
        )

    # 2) Leer archivo
    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel.")

    ws = wb.active

    # 3) Detectar encabezados
    try:
        header_row, header_col = detect_headers(ws)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    max_row = ws.max_row
    max_col = ws.max_column

    # 4) Mapas desde la BD
    country_map, country_ambiguous = build_country_map(db)
    indicator_map, indicator_ambiguous = build_indicator_map(db)

    # 5) Leer encabezados de indicadores (fila)
    indicator_labels: dict[int, str] = {}
    for c in range(1, max_col + 1):
        if c == header_col:
            continue
        cell_value = ws.cell(row=header_row, column=c).value
        if not is_text(cell_value):
            continue
        indicator_labels[c] = normalize_text(str(cell_value))

    # 6) Leer encabezados de países (columna)
    country_labels: dict[int, str] = {}
    for r in range(1, max_row + 1):
        if r == header_row:
            continue
        cell_value = ws.cell(row=r, column=header_col).value
        if not is_text(cell_value):
            continue
        country_labels[r] = normalize_text(str(cell_value))

    if not indicator_labels or not country_labels:
        raise HTTPException(
            status_code=400,
            detail="No se detectaron suficientes indicadores o países en el archivo.",
        )

    # 7) Resolver países e indicadores contra la BD
    row_country_id: dict[int, int] = {}
    col_indicator_id: dict[int, int] = {}
    errors: list[str] = []

    # Países
    for r, norm_name in country_labels.items():
        original = ws.cell(row=r, column=header_col).value
        if norm_name in country_ambiguous:
            errors.append(
                f"Fila {r}: el país '{original}' es ambiguo (coincide con más de un país)."
            )
            continue
        country_obj = country_map.get(norm_name)
        if not country_obj:
            errors.append(
                f"Fila {r}: el país '{original}' no existe en la base de datos."
            )
            continue
        row_country_id[r] = country_obj.id

    # Indicadores
    for c, norm_name in indicator_labels.items():
        original = ws.cell(row=header_row, column=c).value
        if norm_name in indicator_ambiguous:
            errors.append(
                f"Columna {c}: el indicador '{original}' es ambiguo (coincide con más de un indicador)."
            )
            continue
        indicator_obj = indicator_map.get(norm_name)
        if not indicator_obj:
            errors.append(
                f"Columna {c}: el indicador '{original}' no existe en la base de datos."
            )
            continue
        col_indicator_id[c] = indicator_obj.id

    if not row_country_id or not col_indicator_id:
        raise HTTPException(
            status_code=400,
            detail="No se pudo asociar ningún país o indicador del Excel con la base de datos.",
        )

    # 8) Recorrer matriz y hacer upsert a través del repo
    processed = 0

    for r, country_id in row_country_id.items():
        for c, indicator_id in col_indicator_id.items():
            cell = ws.cell(row=r, column=c)
            cell_value = cell.value
            coord = cell.coordinate

            if cell_value is None or cell_value == "":
                continue

            if not is_number(cell_value):
                errors.append(
                    f"Celda {coord}: valor '{cell_value}' no es numérico, se ignora."
                )
                continue

            raw_value = float(cell_value)

            payload = IndicatorValueCreate(
                scenario_id=scenario_id,
                country_id=country_id,
                indicator_id=indicator_id,
                raw_value=raw_value,
            )

            try:
                # usamos tu repo para mantener toda la lógica de normalización igual
                repo.upsert_value(
                    db,
                    payload,
                    user_id=current.id if current else None,
                )
                processed += 1
            except NormalizationError as e:
                errors.append(
                    f"Celda {coord}: error de normalización: {str(e)}"
                )
            except ValueError as e:
                # por ejemplo si el escenario no existe, etc.
                errors.append(
                    f"Celda {coord}: error de datos: {str(e)}"
                )

    return {
        "processed": processed,
        "errors": errors,
    }
